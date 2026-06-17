from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def build_xlsx(path: str, sheets: list[dict]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for spec in sheets or [{"name": "Sheet1", "rows": []}]:
        ws = wb.create_sheet(title=(spec.get("name") or "Sheet")[:31])
        rows = spec.get("rows") or []
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        # Auto-size columns
        if rows:
            ncols = max(len(r) for r in rows)
            for col in range(1, ncols + 1):
                width = 10
                for row in rows[:200]:
                    if col - 1 < len(row) and row[col - 1] is not None:
                        width = max(width, min(60, len(str(row[col - 1])) + 2))
                ws.column_dimensions[get_column_letter(col)].width = width
            ws.freeze_panes = "A2"
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    wb.save(path)
